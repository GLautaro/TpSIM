from openpyxl import Workbook,load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import numpy as np
def Truncate(n, decimals=0):
    multiplier = 10 ** decimals
    return int(n * multiplier) / multiplier

def GenerarExcel(dict_dataframes, nombre_archivo):
    wb = Workbook(nombre_archivo)
    for nombre,df in dict_dataframes.items():
        ws = wb.create_sheet(nombre)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
    wb.save(nombre_archivo)

def ValidarVectorProbabilidades(vector):
    return (sum(vector) == 1)

def CrearDataFrame(cant_rondas, desde, datos, ultima_ronda, probabilidad_exito):
    
    col = ["Iteracion","Resultado"]
    for i in range(cant_rondas):
        num = str(i + desde)
        t1 = "ronda" + num + "_tirada1"
        t2 = "ronda" + num + "_tirada2"
        
        ran_t1 = "rand_" + t1
        rand_t2 = "rand_" + t2

        ac_rond = "ronda_" + num + "_punt_ac"
        ac_total= "ronda_" + num + "_punt_total"
        col = col + [ran_t1, t1, rand_t2, t2, ac_total, ac_rond]
    if cant_rondas + desde < ultima_ronda:
        ultima_ronda = str(ultima_ronda)
        t1 = "ronda" + ultima_ronda + "_tirada1"
        t2 = "ronda" + ultima_ronda + "_tirada2"        
        ran_t1 = "rand_" + t1
        rand_t2 = "rand_" + t2
        ac_rond = "ronda_" + ultima_ronda + "_punt_ac"
        ac_total= "ronda_" + ultima_ronda + "_punt_total"
        col = col + [ran_t1, t1, rand_t2, t2, ac_total, ac_rond]
    
    print(datos[0])
    print(col)    
    df = pd.DataFrame(np.array(datos), columns=col)
    return df, probabilidad_exito

    